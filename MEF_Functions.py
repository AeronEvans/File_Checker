#When implementing checks to a new tab, make sure to add the tab to the logic in the main.py folder.
import openpyxl
ASSET_TYPES = ["Equity","Cash","ETFs","Fixed Income","Other","Real Estate"]


#check that tickers have same issuer in FI files






#to do - Check latest output file
# characteristics check is broken
# smallcapMF asofdate flagged multiple times
# check output 





def classified_search(ws, file):
    for cell in ws['A']:
        if cell.value == "Not Classified":
            print("\t %s worksheet contains not classified entity" %ws.title)
            file.errors.append("%s worksheet contains not classified entity" %ws.title)



def holdings_checks(ws, file):
    AssetType = None
    columns = openpyxl.utils.cell.get_column_interval(1,ws.max_column)
    file.check_value("As of date", ws["A5"].value[-10:])
    file.check_value("Code", ws["A3"].value)
    try:
        for row in map(str, range(10, ws.max_row+1)):
            if ws["A"+row].value == file.details["Code"]:
                for col in columns[20:40]:
                    try:
                        if ws[col+row].value > 100 or ws[col+row].value < 0:
                            file.errors.append("%s worksheet contains errored revenue exposure value" %ws.title)
                    except:
                        pass
            if ws["A"+row].value == None:
                break
            if ws["A"+row].value in ASSET_TYPES:
                AssetType = ws["A"+row].value
                if ws["H"+row].value <0:
                    print("\t %s worksheet contains negative aggregate %s value" %(ws.title, ws["A"+row].value))
                    file.errors.append("%s worksheet contains negative aggregate %s value" %(ws.title, ws["A"+row].value))
                continue
            if AssetType == "Equity":
                for col in columns[7:17]:
                    if ws[col+row].value == None and ws["J"+row].value[0] != ".":
                        print("\t %s missing %s classification data" %(ws["A"+row].value, ws[col+"9"].value))
                        file.errors.append("%s missing %s classification data" %(ws["A"+row].value, ws[col+"9"].value))
        file.check_value("Number of stocks", ws["F10"].value)
        file.check_complete()
    except:
        print("\t Portfolio holdings checks didn't work")


def characteristics_checks(ws, file): #doesn't check as of date due to different date format used in this tab
    if ws["A8"].value == "Number of Stocks":
        file.check_value("Number of stocks", ws["B8"].value)
    elif ws["A9"].value == "Number of Stocks":
        file.check_value("Number of stocks", ws["B9"].value)



    for cell in ws['B']:
        if cell.value == file.details["Code"]:
            title_row = int(cell.coordinate[1])
            
        elif cell.value == "End % Wgt": #look into this? maybe benchmark I can't remember
            file.check_value("Number of stocks", ws["C"+str(int(cell.coordinate[1])+1)].value)
    try:
        for account in ws.iter_cols(min_row=title_row, min_col = 2, max_col=10, max_row=title_row):
            if account[0].value == None or len(account[0].value) == 0:
                break
            else:
                for row in map(str, range(title_row, ws.max_row-8)):
                    if ws[account[0].coordinate[0]+row].value == "" or ws[account[0].coordinate[0]+row].value == None:
                        file.errors.append("%s tab: %s missing characteristics data" %(ws.title,account[0].value))
                        break
    except:
        pass

    file.check_complete()


def market_band_checks(ws, file): #hard coded
    if ws["E9"].value == "Number of Stocks":
        file.check_value("Number of stocks", ws["E11"].value)
        file.check_complete()
    elif ws["C9"].value == "Number of Stocks":
        file.check_value("Number of stocks", ws["C10"].value)
        file.check_complete()
    else:
        print("\t Market band checks didn't work")


def liquidity_30D_checks(ws, file): #wip
    AssetType = None
    columns = openpyxl.utils.cell.get_column_interval(1,ws.max_column)
    Total_Liquidity = 100
    
    for col in columns[5:11]:
        Total_Liquidity += ws[col+"11"].value

    for row in map(str, range(11, ws.max_row+1)):
        if ws["A"+row].value == None:
            break
        if ws["A"+row].value in ASSET_TYPES:
            AssetType = ws["A"+row].value
            continue
        if AssetType == "Equity":
            if ws["D"+row].value == None:
                file.errors.append("%s tab: Equity security missing value for 30D ADV" %ws.title)
    file.check_complete()

def contribution_checks(ws, file):
    file.check_value(ws.title.split("CONTRIBUTION_")[1][:3]+"_start", ws["A5"].value[8:18]) #checking start date of contribution range
    file.check_value("As of date", ws["A5"].value[-10:])
    for row in map(str, range(1, ws.max_row+1)):
        if ws["A"+row].value == file.details["Code"]:
            if float(ws["B"+row].value) < 99.9 or float(ws["B"+row].value) > 100.1:
                file.errors.append("%s tab: Portfolio average weight is not 100" %ws.title)
    file.check_complete()


def attribution_checks(ws, file):
    file.check_value(ws.title.split("ATTR_")[1][:3]+"_start", ws["A5"].value[8:18]) #checking start date of attribution range
    file.check_value("As of date", ws["A5"].value[-10:])
    for row in map(str, range(1, ws.max_row+1)):
        try:
            if ws["A"+row].value == file.details["Code"]:
                if float(ws["B"+row].value) < 99.9 or float(ws["B"+row].value) > 100.1:
                    file.errors.append("%s tab: Portfolio average weight is not 100" %ws.title)
        except:
            file.errors.append("%s tab: Error checking tab" %ws.title)
    file.check_complete()


def lens_checks(ws, file):
    file.check_value("As of date", ws["A5"].value[-10:])
    lens = True
    try:
        for row in map(str, range(10, ws.max_row+1)):
            if ws["A"+row].value == None:
                break
            if ws["A"+row].value[:3] == "LENS":
                Lens = True
            if ws["A"+row].value[:3] == "Non-LENS":
                Lens = False
            elif sum:
                SEC_sum += float(ws["B"+row].value)
        file.check_complete()
    except:
        pass



def liquidity_scenarios_checks(ws, file): #doesn't check as of date due to different date format used in this tab
    SEC_sum = 0
    sum = False
    for row in map(str, range(10, ws.max_row+1)):
            if ws["A"+row].value == None:
                break
            if ws["A"+row].value[:3] == "SEC":
                sum = True
            elif sum:
                SEC_sum += float(ws["B"+row].value)

    if SEC_sum < 99.9 or SEC_sum > 100.1:
        file.errors.append("%s tab: SEC scenario doesn't sum to 100" %ws.title)
    file.check_complete()



def years_to_maturity_checks(ws, file):
    file.check_value("As of date", ws["A5"].value[-10:])
    file.check_complete()


def asset_type_checks(ws, file):
    file.check_value("As of date", ws["A5"].value[-10:])
    try:
        for row in map(str, range(10, ws.max_row+1)):
            if ws["A"+row].value == None:
                break
            if ws["A"+row].value == "Cash":
                if ws["B"+row].value <= 0:
                    file.errors.append("%s tab: Negative aggregate cash position" %ws.title)
    except:
        print("Asset type check didn't work")

    file.check_complete()





#Work in progress

def scenario_checks(ws, file):
    #prop calcs
    pass

def currency_checks(ws, file):
    #check for negative values
    pass

def econ_correlation_checks(ws, file):
    pass

def tracking_error_checks(ws, file):
    pass

def var_checks(ws, file):
    #proprietry bbg calcs
    pass

def active_share_checks(ws, file):
    pass

def trades_checks(ws, file):
    #end weights match current holdings
    pass