import os, time, sys, openpyxl
from openpyxl import load_workbook
from Code.MEF_Class import MonthEndFile
from Code.MEF_Functions import *


#CONSTANTS

DIR_PATH = os.path.dirname(os.path.abspath(__file__))
INPUT_FOLDER = "Month End files"
INPUT_FILES = os.listdir(os.path.join(DIR_PATH,INPUT_FOLDER))



#FUNCTIONS - Program
        
def create_output_file():
    global output
    output = openpyxl.Workbook()
    output_ws = output.active
    a = MonthEndFile("placeholder")
    output_ws.append([i.capitalize() for i in vars(a)])


def class_to_excel(file, output):
    output_ws = output.active
    row = []
    for i in file.__dict__:
        if i != "errors" and i != "details":
            row.append(getattr(file,i))
    details = ''.join([str(x) + ": " + str(y) + " \n" for x, y in zip(file.details.keys(), file.details.values())])[:-2]
    row.append(details)
    row.append(len(file.errors))
    output_ws.append(row)
    for i in range(len(file.errors)):
        output_ws["E"+str(output_ws.max_row + 1)] = file.errors[i-1]
        output_ws.row_dimensions[output_ws.max_row].outlineLevel = 1
        output_ws.row_dimensions[output_ws.max_row].hidden = True
    output_ws.sheet_properties.outlinePr.summaryBelow = False

def save_and_close():
    print('Closed')
    output.save(os.path.join(DIR_PATH,"Output","Summary_Report_"+time.strftime("%d.%m.%y_%H%M")+".xlsx"))
    print("Saved")


#opening files and assigning checks to tabs

def sheet_parser(report):
    start = time.time()
    
    
    global file
    file = MonthEndFile(report.split("\\")[-1])
    print("\n", file.name)
    try:
        wb = load_workbook(os.path.join(DIR_PATH, INPUT_FOLDER, report), keep_links=False, data_only=True)
    except KeyboardInterrupt:
        save_and_close()
        try:
            sys.exit(130)
        except SystemExit:
            os._exit(130)
    except:
        print(os.path.join(DIR_PATH, INPUT_FOLDER, report), " couldn't be opened")
        file.errors.append("Error opening file")
        print("Ensure the only contents of the input folder are raw month end report excel files with xlsx extension, and that none of the input files are open.")
        class_to_excel(file,output)
        return
    print("\t Opening this file took: %ss" % str(time.time()-start)[:4])
    sheets = wb.sheetnames
    file.tabs = len(sheets)


    for sheet in sheets: #the below could definitely be improved with a dictionary + lambda functions. However since there is too much variety in tab naming convention the structure: if * in sheet is used.
        ws = wb.active = wb[sheet]
        classified_search(ws, file)
        
        if "PORTFOLIO" in sheet and "EXCASH" not in sheet:
            holdings_checks(ws, file)
            
        elif "CHARACTERISTICS" in sheet and "SECTOR" not in sheet:
            characteristics_checks(ws, file)

        elif "MARKET_BAND" in sheet:
            market_band_checks(ws, file)

        elif "CONTRIBUTION" in sheet: #gics sectors > Cash
            contribution_checks(ws, file)

        elif "ATTR" in sheet:
            attribution_checks(ws, file)

        elif "LIQUIDITY_SCENARIOS" in sheet:
            liquidity_scenarios_checks(ws, file)

        elif "LENS" in sheet and len(sheet) == 4:
            lens_checks(ws, file)

        elif "LIQUIDITY" and "30D" in sheet:
            liquidity_30D_checks(ws, file)

        elif "ASSET_TYPE" in sheet:
            asset_type_checks(ws, file)
        
        elif "YEARS_TO_MATURITY" in sheet:
            years_to_maturity_checks(ws, file)

    class_to_excel(file, output)


    
#LOGIC


def main():
    create_output_file()
    path_to_watch = os.path.join(DIR_PATH,INPUT_FOLDER)
    before = dict ([(f, None) for f in os.listdir (path_to_watch)])
    print("Month End File checker - press ctrl+c to save output and exit program.")
    """ while True:
        time.sleep (3)
        print("Waiting for files... ", end='\r')
        after = dict ([(f, None) for f in os.listdir (path_to_watch)])
        added = [f for f in after if not f in before]
        before = after
        for report in added:
            sheet_parser(report) """
    for i in INPUT_FILES:
        sheet_parser(os.path.join(DIR_PATH,INPUT_FOLDER,i))
    output.save(os.path.join(DIR_PATH,"Output","Summary_Report_"+time.strftime("%d.%m.%y_%H%M")+".xlsx"))
    print("Saved")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        save_and_close()
        try:
            sys.exit(130)
        except SystemExit:
            os._exit(130)







